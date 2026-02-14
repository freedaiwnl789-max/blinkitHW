#!/usr/bin/env python3
"""
Blinkit Product Web Scraper
- Crawl product URLs from 70000 to 79999
- Extract product name and price
- Filter products containing "hot wheels" keyword
- Export results to Excel
"""

import asyncio
import pandas as pd
import sys
import os
from datetime import datetime
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from src.auth import BlinkitAuth
from src.order.blinkit_order import BlinkitOrder

# Output file - fixed filename for incremental updates
def get_output_filename():
    return Path("blinkit_products.xlsx")

class ProductScraper:
    def __init__(self, start_id=746500, end_id=747000, keyword_filter="hot wheels"):
        """
        Initialize the product scraper
        
        Args:
            start_id: Starting product ID (default 70000)
            end_id: Ending product ID (default 79999)
            keyword_filter: Keyword to filter products (e.g., "hot wheels")
        """
        self.start_id = start_id
        self.end_id = end_id
        self.keyword_filter = keyword_filter.lower()
        self.products = []
        self.auth = None
        self.current_id = start_id
        self.output_file = get_output_filename()

    async def get_product_info(self, page, product_id, verbose=False):
        """Extract product name and price from product page"""
        try:
            url = f"https://blinkit.com/prn/x/prid/{product_id}"
            
            # Navigate to product page
            try:
                response = await page.goto(url, wait_until="domcontentloaded", timeout=15000)
                status = response.status if response else "unknown"
                if verbose:
                    print(f"  [DEBUG] Status: {status}")
            except Exception as e:
                if verbose:
                    print(f"  [DEBUG] Navigation error: {e}")
                return None, None, None
            
            await asyncio.sleep(0.5)
            
            # Check if page has content (not a 404 or error page)
            try:
                page_text = await page.text_content()
                if not page_text or len(page_text) < 50:
                    if verbose:
                        print(f"  [DEBUG] Page text too short: {len(page_text) if page_text else 0} chars")
                    return None, None, None
            except Exception:
                pass
            
            # Extract product name
            product_name = "Unknown"
            try:
                # Try Blinkit's specific selector first
                specific_selector = ".tw-text-500.tw-font-extrabold.tw-line-clamp-50"
                script = f"() => {{ const el = document.querySelector('{specific_selector}'); return el ? (el.innerText || el.textContent) : null; }}"
                text = await page.evaluate(script)
                if text:
                    product_name = text.strip()
                    if verbose:
                        print(f"  [DEBUG] Got name from specific selector: {product_name}")
                else:
                    # Try meta tag
                    meta = await page.evaluate("() => { const m = document.querySelector(\"meta[property='og:title']\"); return m ? m.getAttribute('content') : null; }")
                    if meta:
                        product_name = meta.strip()
                        if verbose:
                            print(f"  [DEBUG] Got name from meta: {product_name}")
                    else:
                        # Try other selectors
                        selectors = ["h1", ".product-title", ".productName", "[data-testid='product-title']", "[class*='ProductName']"]
                        for sel in selectors:
                            script = f"() => {{ const el = document.querySelector('{sel}'); return el ? (el.innerText || el.textContent) : null; }}"
                            text = await page.evaluate(script)
                            if text:
                                product_name = text.strip()
                                if verbose:
                                    print(f"  [DEBUG] Got name from selector {sel}: {product_name}")
                                break
            except Exception as e:
                if verbose:
                    print(f"  [DEBUG] Name extraction error: {e}")
            
            # Extract image URL
            image_url = "N/A"
            try:
                # Extract from ProductCarousel__ImageContainer
                script = """() => {
                    const container = document.querySelector('[class*="ProductCarousel__ImageContainer"]');
                    if (container) {
                        const img = container.querySelector('img');
                        if (img && img.src) {
                            return img.src;
                        }
                    }
                    return null;
                }"""
                img_src = await page.evaluate(script)
                if img_src:
                    image_url = img_src.strip()
                    if verbose:
                        print(f"  [DEBUG] Got image URL: {image_url}")
            except Exception as e:
                if verbose:
                    print(f"  [DEBUG] Image extraction error: {e}")
            
            # Extract price
            price = "N/A"
            try:
                # Common price selectors
                price_selectors = [
                    ".product-price",
                    "[class*='Price']",
                    "[class*='price']",
                    "[data-testid='product-price']"
                ]
                
                for sel in price_selectors:
                    script = f"() => {{ const el = document.querySelector('{sel}'); return el ? (el.innerText || el.textContent) : null; }}"
                    price_text = await page.evaluate(script)
                    if price_text:
                        price = price_text.strip()
                        if verbose:
                            print(f"  [DEBUG] Got price from selector {sel}: {price}")
                        break
                
                # If still not found, try to extract from page text
                if price == "N/A":
                    script = "() => { const text = document.documentElement.innerText; const match = text.match(/₹\\s*(\\d+(?:,\\d+)*)/); return match ? match[0] : null; }"
                    price_match = await page.evaluate(script)
                    if price_match:
                        price = price_match.strip()
                        if verbose:
                            print(f"  [DEBUG] Got price from regex: {price}")
            except Exception as e:
                if verbose:
                    print(f"  [DEBUG] Price extraction error: Price not found")
            
            # Return only if we got a valid product name
            if product_name and product_name != "Unknown":
                return product_name, price, image_url  # image_url will be "N/A" if not found
            else:
                if verbose:
                    print(f"  [DEBUG] No valid product name found")
                return None, None, None
            
        except Exception as e:
            if verbose:
                print(f"  [DEBUG] Unexpected error: {e}")
            return None, None, None

    async def scrape_products(self, headless=True, verbose=False):
        """
        Scrape products from the specified range
        
        Args:
            headless: Whether to run browser in headless mode
            verbose: Whether to show debug logs for each product
        """
        print("\n" + "=" * 70)
        print("BLINKIT PRODUCT SCRAPER")
        print("=" * 70)
        print(f"Scraping product IDs from {self.start_id} to {self.end_id}")
        print(f"Filtering for keyword: '{self.keyword_filter}'")
        print(f"Output file: {self.output_file}")
        print(f"Debug mode: {'ON' if verbose else 'OFF'}")
        print("-" * 70)
        
        try:
            print("Initializing browser...")
            self.auth = BlinkitAuth(headless=headless)
            await self.auth.start_browser()
            
            if not await self.auth.is_logged_in():
                print("[ERROR] Not logged in!")
                await self.auth.close()
                return False
            
            print("[OK] Logged in successfully\n")
            
            total_products = self.end_id - self.start_id + 1
            scraped_count = 0
            filtered_count = 0
            
            for product_id in range(self.start_id, self.end_id + 1):
                self.current_id = product_id
                progress = ((product_id - self.start_id + 1) / total_products) * 100
                
                # Enable verbose for first few products even if not requested
                is_verbose = verbose or (scraped_count + filtered_count < 3)
                
                product_name, price, image_url = await self.get_product_info(self.auth.page, product_id, verbose=is_verbose)
                
                if product_name:
                    scraped_count += 1
                    
                    # Check if product matches keyword filter
                    if self.keyword_filter in product_name.lower():
                        self.products.append({
                            "Product ID": product_id,
                            "Image URL": image_url,
                            "Product Name": product_name,
                            "Price": price
                        })
                        filtered_count += 1
                        print(f"[{progress:.1f}%] [MATCH] PID:{product_id} | {product_name} | {price}")
                    else:
                        print(f"[{progress:.1f}%] [FOUND] PID:{product_id} | {product_name} | {price} (no keyword match)")
                else:
                    print(f"[{progress:.1f}%] [NOT_FOUND] PID:{product_id} (product not available or error)")
                
                # Small delay to avoid overwhelming the server
                await asyncio.sleep(0.3)
            
            print("\n" + "-" * 70)
            print(f"[COMPLETE] Scraping finished")
            print(f"Total products scraped: {scraped_count}")
            print(f"Products with '{self.keyword_filter}': {filtered_count}")
            print("-" * 70)
            
            # Export to Excel
            if self.products:
                self.export_to_excel()
                print(f"\n[SUCCESS] Results exported to {self.output_file}")
            else:
                print(f"\n[INFO] No products found matching '{self.keyword_filter}'")
            
            return True
            
        except KeyboardInterrupt:
            print("\n[STOPPED] Scraping interrupted by user")
            print(f"Scraped {filtered_count} filtered products so far")
            if self.products:
                self.export_to_excel()
            return False
        
        finally:
            if self.auth:
                try:
                    if hasattr(self.auth, 'close'):
                        await self.auth.close()
                    print("[OK] Browser closed")
                except Exception as e:
                    print(f"[DEBUG] Browser close error: {e}")

    def export_to_excel(self):
        """Export scraped products to Excel file, merging with existing data"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            
            # Prepare new data
            df_new = pd.DataFrame(self.products)
            df_new['Product ID'] = df_new['Product ID'].astype(int)
            
            # Reorder columns
            column_order = ['Product ID', 'Product Name', 'Price', 'Image URL']
            df_new = df_new[column_order].set_index('Product ID')
            
            # Load existing data if file exists
            df_merged = None
            if self.output_file.exists():
                try:
                    # Read existing file defensively
                    df_existing = pd.read_excel(self.output_file, dtype=object)

                    # Ensure required columns exist
                    for col in ['Product ID', 'Product Name', 'Price', 'Image URL']:
                        if col not in df_existing.columns:
                            df_existing[col] = pd.NA

                    # Coerce Product ID to numeric, drop invalid rows
                    df_existing['Product ID'] = pd.to_numeric(df_existing['Product ID'], errors='coerce')
                    df_existing = df_existing.dropna(subset=['Product ID'])
                    df_existing['Product ID'] = df_existing['Product ID'].astype(int)
                    df_existing = df_existing.set_index('Product ID')

                    # Merge: new data overwrites old for existing IDs, keep old for IDs not in new data
                    df_merged = df_existing.copy()
                    df_merged.update(df_new)

                    # Add rows that are present in df_new but missing in df_merged
                    new_idx = df_new.index.difference(df_merged.index)
                    if len(new_idx) > 0:
                        df_merged = pd.concat([df_merged, df_new.loc[new_idx]])

                    df_merged = df_merged.sort_index()

                    print(f"[INFO] Merged with {len(df_existing)} existing products")
                    print(f"[INFO] {len(df_new)} new/updated products processed")
                except Exception as e:
                    print(f"[INFO] Could not read existing file, creating new: {e}")
                    df_merged = df_new
            else:
                df_merged = df_new
            
            # Reset index to convert Product ID back to column
            df_merged = df_merged.reset_index()
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = 'Products'
            
            # Write headers
            for col_num, col_name in enumerate(column_order, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = col_name
            
            # Write data and add hyperlinks to Product IDs
            for row_num, row_data in enumerate(df_merged.values, 2):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    
                    # Add hyperlink for Product ID column
                    if col_num == 1:  # Product ID column
                        product_id = int(value)
                        product_url = f"https://blinkit.com/prn/x/prid/{product_id}"
                        cell.value = product_id
                        cell.hyperlink = product_url
                        cell.font = Font(color="0563C1", underline="single")
                    else:
                        cell.value = value
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Format Product ID column as number (no thousand separators)
            for row in ws.iter_rows(min_col=1, max_col=1, min_row=2, max_row=len(df_merged) + 1):
                for cell in row:
                    cell.number_format = '0'
            
            wb.save(str(self.output_file))
            print(f"[INFO] Total products in file: {len(df_merged)}")
            
            return True
        
        except PermissionError as e:
            print(f"[ERROR] Permission denied: {e}")
            print(f"[INFO] Close the Excel file if it's open, or check file permissions")
            return False
        except Exception as e:
            print(f"[ERROR] Failed to export to Excel: {e}")
            return False


async def main():
    """Main entry point"""
    
    print("\n" + "=" * 70)
    print("BLINKIT PRODUCT SCRAPER SETUP")
    print("=" * 70)
    
    # Get user input for product ID range
    try:
        start_input = input("\nEnter starting product ID (default 746500): ").strip()
        start_id = int(start_input) if start_input else 746500
    except ValueError:
        start_id = 746500
    
    try:
        end_input = input("Enter ending product ID (default 747000): ").strip()
        end_id = int(end_input) if end_input else 747000
    except ValueError:
        end_id = 747000
    
    keyword = input("Enter keyword to filter (default 'hot wheels'): ").strip()
    keyword = keyword if keyword else "hot wheels"
    
    headless_input = input("Run browser in headless mode? (y/N): ").strip().lower()
    headless = headless_input in ('y', 'yes')
    
    debug_input = input("Enable debug/verbose mode? (y/N): ").strip().lower()
    debug = debug_input in ('y', 'yes')
    
    print(f"\n[INFO] Starting product ID: {start_id}")
    print(f"[INFO] Ending product ID: {end_id}")
    print(f"[INFO] Keyword filter: '{keyword}'")
    print(f"[INFO] Headless mode: {headless}")
    print(f"[INFO] Debug mode: {debug}")
    
    # Validate range
    if start_id >= end_id:
        print("[ERROR] Starting ID must be less than ending ID")
        return
    
    if end_id - start_id > 10000:
        confirm = input(f"\n[WARNING] This will scrape {end_id - start_id + 1} products. Continue? (y/N): ").strip().lower()
        if confirm not in ('y', 'yes'):
            print("[CANCELLED] Scraping cancelled")
            return
    
    # Start scraping
    scraper = ProductScraper(start_id, end_id, keyword)
    success = await scraper.scrape_products(headless=headless, verbose=debug)
    
    if success:
        print("\n[SUCCESS] Scraping completed successfully!")
    else:
        print("\n[INFO] Scraping stopped.")


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except Exception as e:
        print(f"[ERROR] Fatal error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
